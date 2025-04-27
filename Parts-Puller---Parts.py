import requests
from lxml import html
import json
import logging
from datetime import datetime
import pytz
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(
    filename='price_puller.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Set timezone to EST
est = pytz.timezone('US/Eastern')

def get_product_info(url, product_name):
    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "user-agent": "Mozilla/5.0"
    }

    try:
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            logging.warning(f"❌ Failed to fetch page for {product_name} | Status: {response.status_code}")
            return None

        tree = html.fromstring(response.content)
        script_content = tree.xpath('//script[@type="application/ld+json"]/text()')
        if not script_content:
            logging.warning(f"⚠️ No JSON-LD found for {product_name}")
            return None

        data = json.loads(script_content[0])

        # Try to get the name from the site if the given product_name is missing or generic
        name_from_site = data.get('name', product_name) or product_name

        offers = data.get('offers', {})
        if isinstance(offers, list):
            offers = offers[0]

        try:
            price = float(offers.get('price', "0"))
        except (ValueError, TypeError):
            price = None

        sku = offers.get('sku', "N/A")

        return {
            "Name": name_from_site,
            "SKU": sku,
            "Price": price,
            "URL": url
        }

    except Exception as e:
        logging.error(f"🚨 Error fetching data for {product_name}: {e}")
        return None

def merge_price_label_row(ws, start_col, product_count):
    from_col = get_column_letter(start_col)
    to_col = get_column_letter(start_col + product_count - 1)
    ws.merge_cells(f'{from_col}3:{to_col}3')
    ws[f'{from_col}3'] = "Price (USD)"
    ws[f'{from_col}3'].font = Font(bold=True)
    ws[f'{from_col}3'].alignment = Alignment(horizontal='center')

def write_to_excel(section_name, product_results, today_str, file_path='CarParts_Pricing.xlsx'):
    base_date = datetime(2025, 4, 24)
    today_date = datetime.strptime(today_str, "%m/%d/%Y")
    days_since = (today_date - base_date).days
    row_index = 4 + days_since

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        # Remove default Sheet if it exists
        default_sheet = wb.active
        if default_sheet.title == "Sheet":
            wb.remove(default_sheet)

    if section_name in wb.sheetnames:
        ws = wb[section_name]
    else:
        ws = wb.create_sheet(section_name)
        ws['A1'] = 'Name'
        ws['A2'] = 'Part # / SKU'
        ws['A3'] = 'Date'
        for r in range(1, 4):
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')

    start_col = 2
    for idx, product in enumerate(product_results):
        name = product["Name"]
        sku = product["SKU"]
        price = product["Price"]
        url = product["URL"]
        col = start_col + idx

        # Name with hyperlink
        name_cell = ws.cell(row=1, column=col)
        name_cell.value = name
        if url:
            name_cell.hyperlink = url
            name_cell.style = "Hyperlink"

        # SKU
        ws.cell(row=2, column=col).value = sku

        # Price with color-coded change detection
        price_cell = ws.cell(row=row_index, column=col)
        price_cell.value = price
        price_cell.number_format = '"$"#,##0.00'

        previous_price_cell = ws.cell(row=row_index - 1, column=col)
        previous_price = previous_price_cell.value

        if previous_price is not None and isinstance(previous_price, (int, float)) and isinstance(price, (int, float)):
            if price > previous_price:
                color = "FF0000"  # red
            elif price < previous_price:
                color = "00B050"  # green
            else:
                color = "000000"  # black
            price_cell.font = Font(color=color)
        else:
            price_cell.font = Font(color="000000")  # default black

    # Merge and center Price (USD)
    merge_price_label_row(ws, start_col, len(product_results))

    # Insert date in A column
    ws.cell(row=row_index, column=1).value = today_date
    ws.cell(row=row_index, column=1).number_format = 'mm/dd/yyyy'

    # Auto-fit columns based on Row 1 content
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            width = len(str(cell_value)) + 2
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width

    try:
        wb.save(file_path)
        logging.info(f"✅ Excel updated for section: {section_name}")
    except PermissionError:
        print("❌ Excel File open — please close the file and try again.")
        logging.error("❌ Excel File open — save failed.")

def main():
    today_str = datetime.now(est).strftime("%m/%d/%Y")

    car_parts = {
        "B8.5 S4": [
            ("https://www.fcpeuro.com/products/audi-engine-mount-kit-034motorsport-0345090050sdkt-kit-0345090050sdkt"),
            ("https://www.fcpeuro.com/products/audi-cv-axle-assembly-front-gkn-8k0407271aj"),
            ("https://www.fcpeuro.com/products/audi-cv-boot-kit-gkn-8k0498201b"),
            ("https://www.fcpeuro.com/products/porsche-cabin-air-filter-corteco-8k0819439b"),
            ("https://www.fcpeuro.com/products/audi-a-c-compressor-denso-8t0260805g"),
            ("https://www.fcpeuro.com/products/audi-a-c-line-o-ring-r8-a5-quattro-s5-a5-vag-4e0260749a"),
            ("https://www.fcpeuro.com/products/audi-drive-belt-tensioner-febi-43784"),
            ("https://www.fcpeuro.com/products/vw-audi-accessory-drive-belt-6k2330"),
            ("https://www.fcpeuro.com/products/audi-high-pressure-fuel-pump-kit-hitachi-hpp0009kt2"),
            ("https://www.fcpeuro.com/products/audi-timing-chain-kit-iwis-06e109465kt"),
            ("https://www.fcpeuro.com/products/audi-ignition-coil-kit-denso-06e905115fkt10"),
            ("https://www.fcpeuro.com/products/audi-spark-plug-kit-ngk-101905631gkt"),
            ("https://www.fcpeuro.com/products/audi-fuel-pump-assembly-vdo-8k0919051aj"),
            ("https://www.fcpeuro.com/products/audi-wheel-bearing-and-hub-kit-fag-febi-7136109700kt1"),
            ("https://www.fcpeuro.com/products/audi-vw-dsg-transmission-service-kit-genuine-audi-vw-516089"),
            ("https://www.fcpeuro.com/products/audi-engine-coolant-bypass-hose-seal-genuine-audi-07l103121r"),
            ("https://www.fcpeuro.com/products/audi-vw-pcv-valve-hose-06e103213"),
            ("https://www.fcpeuro.com/products/audi-porsche-vw-engine-water-pump-vne-4019000"),
            ("https://www.fcpeuro.com/products/audi-vw-engine-coolant-thermostat-vag-06e121111al"),
            ("https://www.fcpeuro.com/products/audi-crankcase-breather-hose-vne-06e103207ap"),
            ("https://www.fcpeuro.com/products/clone-audi-control-arm-kit-front-8-piece-a4-quattro-a5-quattro-s4-s5-lemforder-b8cakitoe"),
            ("https://www.fcpeuro.com/products/audi-control-arm-hardware-kit-genuine-vw-audi-n10425302"),
            ("https://www.fcpeuro.com/products/audi-shock-mount-kit-sachs-8k0412377ckt"),
            ("https://www.fcpeuro.com/products/audi-supercharger-belt-kit-continental-7pk1270kt"),
            ("https://www.fcpeuro.com/products/audi-air-filter-a5-quattro-s5-q5-s4-a4-quattro-a5-e675ld157")
            ],

        "E46 M3": [
            ("https://www.fcpeuro.com/products/bmw-clutch-flywheel-z4-m3-dmf050", "BMW Z4 M3 Clutch Flywheel"),
            ("https://www.fcpeuro.com/products/bmw-clutch-kit-m3-e46-luk-21212282667", "BMW E46 M3 Clutch Kit"),
            ("https://www.fcpeuro.com/products/bmw-clutch-slave-cylinder-oem-21526785966", "BMW Clutch Slave Cylinder OEM"),
            ("https://www.fcpeuro.com/products/bmw-clutch-pilot-bearing-11211720310", "BMW Clutch Pilot Bearing"),
            ("https://www.fcpeuro.com/products/bmw-differential-cover-kit-33112282482kt", "BMW Differential Cover Kit"),
            ("https://www.fcpeuro.com/products/bmw-oil-change-kit-e46-m3-z3m-z4m-11427833769kt2", "BMW Oil Change Kit"),
            ("https://www.fcpeuro.com/products/bmw-cabin-air-filter-64319257504-1", "BMW Cabin Air Filter"),
            ("https://www.fcpeuro.com/products/bmw-wiper-blade-set-bosch-61610037009", "BMW Wiper Blade Set"),
            ("https://www.fcpeuro.com/products/jectron-fuel-injection-cleaner-liqui-moly-lm2007", "Fuel Injection Cleaner"),
            ("https://www.fcpeuro.com/products/bmw-emblem-rear-trunk-lid-51148219237", "Rear Trunk Lid Emblem"),
            ("https://www.fcpeuro.com/products/fuel-system-cleaner-500ml-can-liqui-moly-lm2030", "Fuel System Cleaner"),
            ("https://www.fcpeuro.com/products/bmw-oxygen-sensor-front-upper-m3-z3-bosch-13949"),
            ("https://www.fcpeuro.com/products/bmw-oxygen-sensor-rear-upper-m3-z3-bosch-13952"),
            ("https://www.fcpeuro.com/products/bmw-oxygen-sensor-front-lower-m3-z3-z4-bosch-15138"),
            ("https://www.fcpeuro.com/products/bmw-oxygen-sensor-rear-lower-m3-z3-z4-bosch-13951"),
            ("https://www.fcpeuro.com/products/bmw-egt-sensor-oem-supplier-11787831624")
        ],

        "E92 M3": [
            ("https://www.fcpeuro.com/products/bmw-brake-kit-front-and-rear-e9xm3brakekit2"),
            ("https://www.fcpeuro.com/products/bmw-10-piece-control-arm-kit-rear-1m-m3-e9xm310piecerearoe"),
            ("https://www.fcpeuro.com/products/bmw-bushing-kit-rear-m-spec-33326763092kit"),
            ("https://www.fcpeuro.com/products/bmw-engine-mount-kit-11812283798kt"),
            ("https://www.fcpeuro.com/products/bmw-drive-shaft-end-bushing-26117526611"),
            ("https://www.fcpeuro.com/products/bmw-transmission-mount-oe-supplier-22312283285"),
            ("https://www.fcpeuro.com/products/bmw-10-piece-control-arm-kit-front-1m-m3-e9xm310piecekitoe"),
            ("https://www.fcpeuro.com/products/bmw-control-arm-ball-joint-kit-lemforder-33326792553kt"),
            ("https://www.fcpeuro.com/products/bmw-wheel-hub-assembly-kit-31222282670kt"),
            ("https://www.fcpeuro.com/products/bmw-spark-plug-set-m3-e90-e92-e93-m3sparkplugset"),
            ("https://www.fcpeuro.com/products/bmw-ambient-temperature-sensor-repair-kit-65816936953kt1"),
            ("https://www.fcpeuro.com/products/bmw-intake-air-temperature-sensor-hella-13627792203"),
            ("https://www.fcpeuro.com/products/bmw-carbon-activated-cabin-air-filter-set-e90-e92-e93-m3-corteco-64319159606"),
            ("https://www.fcpeuro.com/products/bmw-a-c-compressor-m3-m5-m6-471-1559"),
            ("https://www.fcpeuro.com/products/bmw-throttle-actuator-gear-repair-kit-odometer-gears"),
            ("https://www.fcpeuro.com/products/bmw-throttle-body-actuator-vdo-13627838085"),
            ("https://www.fcpeuro.com/products/bmw-drive-belt-tensioner-pulley-ina-11287838194"),
            ("https://www.fcpeuro.com/products/bmw-black-carbon-intake-plenum-eventuri-eve-e9x-cf-plm"),
            ("https://www.fcpeuro.com/products/bmw-black-carbon-intake-eventuri-eve-e9x-cf-int"),
            ("https://www.fcpeuro.com/products/bmw-gloss-carbon-air-duct-suction-hood-eventuri-eve-e9x-cf-dct"),
            ("https://www.fcpeuro.com/products/bmw-expansion-tank-kit-17112283500kt"),
            ("https://www.fcpeuro.com/products/bmw-water-pump-and-thermostat-kit-11517838201kt"),
            ("https://www.fcpeuro.com/products/bmw-thermostat-housing-11537838480"),
            ("https://www.fcpeuro.com/products/bmw-oil-change-kit-e90-e92-e93-m3-liqui-moly-mahle-11427837997kt2-lm"),
            ("https://www.fcpeuro.com/products/bmw-oil-filter-housing-gasket-kit-11427838276kt1"),
            ("https://www.fcpeuro.com/products/bmw-oil-pan-gasket-elring-11137841085"),
            ("https://www.fcpeuro.com/products/bmw-isa-screw-with-washer-genuine-bmw-07129905599"),
            ("https://www.fcpeuro.com/products/bmw-isa-screw-with-washer-m6x75znniv-si-07129905600"),
            ("https://www.fcpeuro.com/products/bmw-isa-screw-with-washer-07129905537"),
            ("https://www.fcpeuro.com/products/bmw-engine-oil-pump-pickup-tube-gasket-m3-11417839832"),
            ("https://www.fcpeuro.com/products/bmw-engine-oil-pump-pickup-tube-gasket-m3-11417839833"),
            ("https://www.fcpeuro.com/products/bmw-torx-bolt-with-washer-m6x16-u1-07129904819"),
            ("https://www.fcpeuro.com/products/bmw-screw-with-washer-m12x15x9010-9-31106769907"),
            ("https://www.fcpeuro.com/products/bmw-hexagon-screw-with-flange-m12x15x145-31106769908"),
            ("https://www.fcpeuro.com/products/bmw-hexagon-screw-with-flange-m12x15x5310-9-31106767497"),
            ("https://www.fcpeuro.com/products/bmw-hex-screw-with-collar-m10x3310-9-33326768354"),
            ("https://www.fcpeuro.com/products/bmw-hex-bolt-with-washer-m12x1-5x115-8-8-31106763928"),
            ("https://www.fcpeuro.com/products/bmw-locking-nut-m12x1510-zns3-31106767496"),
            ("https://www.fcpeuro.com/products/bmw-self-locking-hex-nut-33326760668"),
            ("https://www.fcpeuro.com/products/bmw-combination-nut-33306787062"),
            ("https://www.fcpeuro.com/products/bmw-steering-torx-bolt-32306778609"),
            ("https://www.fcpeuro.com/products/bmw-gasket-ring-07119906464"),
            ("https://www.fcpeuro.com/products/bmw-gasket-ring-07119906463"),
            ("https://www.fcpeuro.com/products/bmw-oil-filter-m3-ox254d3"),
            ("https://www.fcpeuro.com/products/bmw-rod-bearing-replacement-kit-11247841703kt"),
            ("https://www.fcpeuro.com/products/bmw-drive-belt-kit-11287841529kt1"),
            ("https://www.fcpeuro.com/products/bmw-spark-plug-tube-uro-parts-11127835170"),
            ("https://www.fcpeuro.com/products/bmw-valve-cover-gasket-kit-11127838271kt1"),
            ("https://www.fcpeuro.com/products/bmw-engine-valve-cover-set-oem-grey-nrw-design-s65vcoeg"),
            ("https://www.fcpeuro.com/products/bmw-crankshaft-sensor-hella-13627525015"),
            ("https://www.fcpeuro.com/products/bmw-engine-camshaft-position-sensor-13627837904"),
            ("https://www.fcpeuro.com/products/bmw-engine-crankshaft-main-bearing-m5-m6-m3-11217841606"),
            ("https://www.fcpeuro.com/products/bmw-engine-crankshaft-main-bearing-m5-m6-m3-11217841610"),
            ("https://www.fcpeuro.com/products/bmw-engine-crankshaft-main-bearing-m5-m6-m3-11217841611"),
            ("https://www.fcpeuro.com/products/bmw-engine-crankshaft-main-bearing-m5-m6-m3-11217841609"),
            ("https://www.fcpeuro.com/products/bmw-engine-crankshaft-main-bearing-cap-bolt-m6-m5-11117838986"),
            ("https://www.fcpeuro.com/products/bmw-engine-crankshaft-main-bearing-cap-bolt-m5-m6-m3-11117834141"),
            ("https://www.fcpeuro.com/products/bmw-starter-bosch-12417843530"),
            ("https://www.fcpeuro.com/products/bmw-muffler-hanger-kit-18201401797kt2"),
            ("https://www.fcpeuro.com/products/bmw-wiper-blade-set-aerotwin-e92-3397007579"),
            ("https://www.fcpeuro.com/products/bmw-drive-shaft-flex-joint-giubo-febi-26112282573"),
            ("https://www.fcpeuro.com/products/bmw-driveshaft-center-support-febi-26122283046"),
            ("https://www.fcpeuro.com/products/bmw-taillight-ulo-63217251959"),
            ("https://www.fcpeuro.com/products/bmw-taillight-ulo-63217251960"),
            ("https://www.fcpeuro.com/products/bmw-abs-wheel-speed-sensor-rear-ate-34526785022"),
            ("https://www.fcpeuro.com/products/bmw-stainless-steel-brake-line-set-m3-31053"),
            ("https://www.fcpeuro.com/products/bmw-stainless-steel-brake-line-kit-stop-tech-950-34019"),
            ("https://www.fcpeuro.com/products/bmw-connecting-rod-bolt-set-201-6001"),
            ("https://www.fcpeuro.com/products/bmw-radiator-60786a"),
            ("https://www.fcpeuro.com/products/bmw-floor-mat-w-heelpad-m3-embr-black-e92-82110439366")
        ],

        "E39 M5": [
            ("https://www.fcpeuro.com/products/bmw-control-arm-kit-10-piece-540i-m5-e39-540e3910piece-l"),
            ("https://www.fcpeuro.com/products/bmw-accessory-drive-belt-kit-continental-7pk1629kt"),
            ("https://www.fcpeuro.com/products/bmw-drive-shaft-flex-joint-kit-26112228781kt"),
            ("https://www.fcpeuro.com/products/bmw-subframe-bushing-kit-lemforder-33311091422kt"),
            ("https://www.fcpeuro.com/products/bmw-hydraulic-valve-lifter-set-ina-11321748884"),
            ("https://www.fcpeuro.com/products/bmw-engine-mount-kit-22111092895kt"),
            ("https://www.fcpeuro.com/products/bmw-differential-mount-kit-lemforder-33171093008kt"),
            ("https://www.fcpeuro.com/products/bmw-differential-service-kit-e46-e90-e92-e93-genuine-bmw-33117525064kt6"),
            ("https://www.fcpeuro.com/products/bmw-ambient-temperature-sensor-repair-kit-65816936953kt1"),
            ("https://www.fcpeuro.com/products/1696011295-bmw-spark-plug-kit-set-of-6-12120037607kt13"),
            ("https://www.fcpeuro.com/products/bmw-ignition-coils-set-of-8-12131748017brx8"),
            ("https://www.fcpeuro.com/products/bmw-cabin-air-filter-set-528i-540i-e39-042-2012"),
            ("https://www.fcpeuro.com/products/bmw-oil-filter-all-paper-11421745390"),
            ("https://www.fcpeuro.com/products/bmw-air-filter-oem-13721736675"),
            ("https://www.fcpeuro.com/products/bmw-fuel-filter-m5-z8-xengst-13321407299"),
            ("https://www.fcpeuro.com/products/10w-60-motor-oil-full-synthetic-5-litre-lm2024"),
            ("https://www.fcpeuro.com/products/10w-60-motor-oil-full-synthetic-1-litre-lm2068"),
            ("https://www.fcpeuro.com/products/bmw-a-c-compressor-540i-740i-740il-m5-z8-471-1121"),
            ("https://www.fcpeuro.com/products/bmw-brake-hose-rear-34321162616"),
            ("https://www.fcpeuro.com/products/bmw-clutch-master-cylinder-fte-21526879477"),
            ("https://www.fcpeuro.com/products/bmw-brake-master-cylinder-n-a-34311165544oe"),
            ("https://www.fcpeuro.com/products/bmw-clutch-kit-m5-z8-03-042"),
            ("https://www.fcpeuro.com/products/bmw-clutch-flywheel-m5-21212229190"),
            ("https://www.fcpeuro.com/products/bmw-cooling-fan-clutch-m5-z8-behr-11527830486"),
            ("https://www.fcpeuro.com/products/bmw-clutch-fork-pivot-pin-aga-21511223328"),
            ("https://www.fcpeuro.com/products/bmw-clutch-release-arm-uro-parts-21511223302"),
            ("https://www.fcpeuro.com/products/bmw-auxiliary-fan-switch-fae-13621433077"),
            ("https://www.fcpeuro.com/products/bmw-radiator-m5-z8-e39-e52-behr-17111436062"),
            ("https://www.fcpeuro.com/products/bmw-water-pump-laso-11511407113"),
            ("https://www.fcpeuro.com/products/bmw-coolant-thermostat-m5-behr-11537835558"),
            ("https://www.fcpeuro.com/products/bmw-differential-service-kit-e46-e90-e92-e93-33117525064kt1"),
            ("https://www.fcpeuro.com/products/bmw-manual-trans-service-kit-e36-e39-e46-23117527440kt1"),
            ("https://www.fcpeuro.com/products/bmw-ac-drive-belt-kit-continental-5pk980kt1"),
            ("https://www.fcpeuro.com/products/bmw-s62-timing-chain-kit-s62timingkit"),
            ("https://www.fcpeuro.com/products/bmw-pcv-breather-system-11151406788kt"),
            ("https://www.fcpeuro.com/products/bmw-wheel-center-cap-pack-of-4-genuine-bmw-36136783536kt"),
            ("https://www.fcpeuro.com/products/bmw-right-headlight-white-turn-indicator-63126902518"),
            ("https://www.fcpeuro.com/products/bmw-left-headlight-white-turn-indicator-63126902517"),
            ("https://www.fcpeuro.com/products/bmw-right-rear-light-white-turn-indicator-63216902530"),
            ("https://www.fcpeuro.com/products/bmw-left-rear-light-white-turn-indicator-63216902529"),
            ("https://www.fcpeuro.com/products/bmw-m-steering-wheel-32342282020"),
            ("https://www.fcpeuro.com/products/bmw-expansion-tank-m5-z8-17112229114"),
            ("https://www.fcpeuro.com/products/bmw-expansion-tank-cap-17111742231"),
            ("https://www.fcpeuro.com/products/mini-fuel-tank-cap-cooper-16117222391"),
            ("https://www.fcpeuro.com/products/bmw-final-stage-unit-acm-64116923204"),
            ("https://www.fcpeuro.com/products/bmw-72-56mm-wheel-hub-conversion-kit-e39-31226765601kt"),
            ("https://www.fcpeuro.com/products/bmw-s62-comprehensive-rod-bearing-kit-11241407493kt")
        ],

        "W204 C63 AMG": [
            ("https://www.fcpeuro.com/products/mercedes-m156-camshaft-replacement-kit-156050p"),
            ("https://www.fcpeuro.com/products/mercedes-hydraulic-lifter-replacement-kit-ina-1560500225"),
            ("https://www.fcpeuro.com/products/mercedes-spark-plug-ilzkar7a10"),
            ("https://www.fcpeuro.com/products/mercedes-benz-air-filter-r63-amg-ml63-amg-e63-amg-cls63-042-1731"),
            ("https://www.fcpeuro.com/products/mercedes-direct-ignition-coil-delphi-1569064400"),
            ("https://www.fcpeuro.com/products/mercedes-engine-camshaft-follower-set-ina-1560500225"),
            ("https://www.fcpeuro.com/products/mercedes-control-arm-kit-lemforder-212350"),
            ("https://www.fcpeuro.com/products/mercedes-fuel-injector-kit-bosch-1560780023"),
            ("https://www.fcpeuro.com/products/mercedes-blower-motor-repair-kit-mahle-behr-2128200708"),
            ("https://www.fcpeuro.com/products/mercedes-engine-mount-kit-oe-supplier-2042405117"),
            ("https://www.fcpeuro.com/products/mercedes-drive-belt-kit-m156-genuine-mercedes-m156beltkit"),
            ("https://www.fcpeuro.com/products/mercedes-limited-slip-differential-service-kit-castrol-001989520310"),
            ("https://www.fcpeuro.com/products/mercedes-sway-bar-link-kit-lemforder-2043201"),
            ("https://www.fcpeuro.com/products/mercedes-722-9-transmission-service-kit-liqui-moly-7229late"),
            ("https://www.fcpeuro.com/products/mercedes-a-c-compressor-denso-0022303111"),
            ("https://www.fcpeuro.com/products/mercedes-air-filter-cover-kit-genuine-mercedes-156094"),
            ("https://www.fcpeuro.com/products/mercedes-brake-kit-front-c63-amg-zimmerman-w204amgfbk1"),
            ("https://www.fcpeuro.com/products/mercedes-brake-kit-zimmermann-2044230412"),
            ("https://www.fcpeuro.com/products/mercedes-radiator-1975000003"),
            ("https://www.fcpeuro.com/products/mercedes-coolant-expansion-tank-mahle-behr-2045000949"),
            ("https://www.fcpeuro.com/products/mercedes-engine-water-pump-mer-1562000601"),
            ("https://www.fcpeuro.com/products/mercedes-drive-shaft-flex-joint-kit-febi-0004110600kt"),
            ("https://www.fcpeuro.com/products/mercedes-wheel-bearing-rear-c300-we60394"),
            ("https://www.fcpeuro.com/products/mercedes-wheel-bearing-c300-we60395"),
            ("https://www.fcpeuro.com/products/mercedes-bolt-febi-0009907503-feb-40617"),
            ("https://www.fcpeuro.com/products/mercedes-wheel-hub-nut-0003531373"),
            ("https://www.fcpeuro.com/products/mercedes-axle-hub-2113570508"),
            ("https://www.fcpeuro.com/products/mercedes-wheel-hub-meyle-0147520009"),
            ("https://www.fcpeuro.com/products/mercedes-drive-shaft-center-support-meyle-0140410073s"),
            ("https://www.fcpeuro.com/products/mercedes-power-steering-pump-bosch-006466880180"),
            ("https://www.fcpeuro.com/products/mercedes-power-steering-reservoir-replacement-kit-genuine-mercedes-0004669502"),
            ("https://www.fcpeuro.com/products/mercedes-steering-tie-rod-assembly-febi-44692"),
            ("https://www.fcpeuro.com/products/mercedes-steering-tie-rod-assembly-febi-44691"),
            ("https://www.fcpeuro.com/products/mercedes-control-arm-kit-lemforder-212350"),
            ("https://www.fcpeuro.com/products/mercedes-auto-trans-overhaul-kit-transtec-2694"),
            ("https://www.fcpeuro.com/products/mercedes-auto-trans-mount-e63-amg-mer-2122401018"),
            ("https://www.fcpeuro.com/products/mercedes-steering-rack-boot-febi-2044630296"),
            ("https://www.fcpeuro.com/products/mercedes-thermostat-1562030475"),
        ],

        "991 GT3": [
            ("https://www.fcpeuro.com/products/porsche-center-lock-wheel-nut-kit-genuine-porsche-99136108190kt4"),
            ("https://www.fcpeuro.com/products/porsche-wheel-bearing-kit-aftermarket-95834190100"),
            ("https://www.fcpeuro.com/products/bosch-windshield-wiper-set-bmw-porsche-land-rover-bosch-3397007697"),
            ("https://www.fcpeuro.com/products/porsche-engine-oil-change-kit-5w-40-liqui-moly-991gt3oilkt7"),
            ("https://www.fcpeuro.com/products/volvo-mercedes-benz-serpentine-belt-slk230-960-s90-v90-6pk1755"),
            ("https://www.fcpeuro.com/products/porsche-air-filter-hengst-e1722l"),
            ("https://www.fcpeuro.com/products/porsche-cabin-air-filter-hengst-e3940lb"),
            ("https://www.fcpeuro.com/products/porsche-cabin-air-filter-hengst-e3945lb"),
            ("https://www.fcpeuro.com/products/porsche-brake-kit-sebro-991gt3brkt8"),
            ("https://www.fcpeuro.com/products/porsche-center-lock-wheel-socket-cta-5016"),
            ("https://www.fcpeuro.com/products/porsche-center-lock-wheel-center-cap-removal-tool-genuine-porsche-99136107900"),
            ("https://www.fcpeuro.com/products/porsche-dual-clutch-transmission-service-kit-liqui-moly-9g132102500kt1")
        ],
        "9th Gen Accord": [
            ("https://www.fcpeuro.com/products/headlight-light-bulb-upgrade-philips-crystalvision-h11cvps2", "Headlight Upgrade Bulb")
        ]
    }

    for section, items in car_parts.items():
        print(f"\n📦 Processing section: {section}")
        product_data = []

        for item in items:
            if isinstance(item, tuple):
                url, name = item
            else:
                url = item
                name = "Unnamed Product"

            data = get_product_info(url, name)
            if data:
                product_data.append(data)

        if product_data:
            write_to_excel(section, product_data, today_str)

if __name__ == '__main__':
    main()

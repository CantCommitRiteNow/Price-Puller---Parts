# price_scraper/utils.py

import os
from datetime import datetime
import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

est = pytz.timezone('US/Eastern')

def merge_price_label_row(ws, start_col, product_count):
    from_col = get_column_letter(start_col)
    to_col = get_column_letter(start_col + product_count - 1)
    ws.merge_cells(f'{from_col}3:{to_col}3')
    ws[f'{from_col}3'] = "Price (USD)"
    ws[f'{from_col}3'].font = Font(bold=True)
    ws[f'{from_col}3'].alignment = Alignment(horizontal='center')

def write_to_excel(section_name, product_results, today_str, file_path='CarParts_Pricing.xlsx'):
    base_date = datetime(2025, 4, 24)
    today_date = datetime.strptime(today_str, "%m/%d/%Y")
    days_since = (today_date - base_date).days
    row_index = 4 + days_since

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet.title == "Sheet":
            wb.remove(default_sheet)

    if section_name in wb.sheetnames:
        ws = wb[section_name]
    else:
        ws = wb.create_sheet(section_name)
        ws['A1'] = 'Name'
        ws['A2'] = 'Part # / SKU'
        ws['A3'] = 'Date'
        for r in range(1, 4):
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')

    start_col = 2
    for idx, product in enumerate(product_results):
        name = product["Name"]
        sku = product["SKU"]
        price = product["Price"]
        url = product["URL"]
        col = start_col + idx

        name_cell = ws.cell(row=1, column=col)
        name_cell.value = name
        if url:
            name_cell.hyperlink = url
            name_cell.style = "Hyperlink"

        ws.cell(row=2, column=col).value = sku

        price_cell = ws.cell(row=row_index, column=col)
        price_cell.value = price
        price_cell.number_format = '"$"#,##0.00'

        previous_price_cell = ws.cell(row=row_index - 1, column=col)
        previous_price = previous_price_cell.value

        if previous_price is not None and isinstance(previous_price, (int, float)) and isinstance(price, (int, float)):
            if price > previous_price:
                color = "FF0000"
            elif price < previous_price:
                color = "00B050"
            else:
                color = "000000"
            price_cell.font = Font(color=color)
        else:
            price_cell.font = Font(color="000000")

    merge_price_label_row(ws, start_col, len(product_results))

    ws.cell(row=row_index, column=1).value = today_date
    ws.cell(row=row_index, column=1).number_format = 'mm/dd/yyyy'

    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            width = len(str(cell_value)) + 2
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width

    try:
        wb.save(file_path)
        print(f"✅ Saved pricing to {file_path}")
    except PermissionError:
        print("❌ Excel file is open! Please close it and try again.")

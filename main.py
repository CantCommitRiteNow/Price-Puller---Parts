import os
import requests
import json
import logging
import shutil
from datetime import datetime
from lxml import html
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pytz

# Setup logging
logging.basicConfig(
    filename='price_puller.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Set timezone
est = pytz.timezone('US/Eastern')

# Backup input_links.txt before starting
def backup_input_file():
    backup_dir = "backups"
    os.makedirs(backup_dir, exist_ok=True)

    if os.path.exists('input_links.txt'):
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = os.path.join(backup_dir, f"input_links_backup_{now}.txt")
        shutil.copy('input_links.txt', backup_name)
        logging.info(f"🛡️ Backup created: {backup_name}")

# Scrape product info
def get_product_info(url, product_name):
    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "user-agent": "Mozilla/5.0"
    }

    try:
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            logging.warning(f"❌ Failed to fetch page for {product_name} | Status: {response.status_code}")
            return None

        tree = html.fromstring(response.content)
        script_content = tree.xpath('//script[@type="application/ld+json"]/text()')
        if not script_content:
            logging.warning(f"⚠️ No JSON-LD found for {product_name}")
            return None

        data = json.loads(script_content[0])

        name_from_site = data.get('name', product_name) or product_name
        offers = data.get('offers', {})
        if isinstance(offers, list):
            offers = offers[0]

        try:
            price = float(offers.get('price', "0"))
        except (ValueError, TypeError):
            price = None

        sku = offers.get('sku', "N/A")

        return {
            "Name": name_from_site,
            "SKU": sku,
            "Price": price,
            "URL": url
        }

    except Exception as e:
        logging.error(f"🚨 Error fetching data for {product_name}: {e}")
        return None

# Merge "Price (USD)" header row
def merge_price_label_row(ws, start_col, product_count):
    from_col = get_column_letter(start_col)
    to_col = get_column_letter(start_col + product_count - 1)
    ws.merge_cells(f'{from_col}3:{to_col}3')
    ws[f'{from_col}3'] = "Price (USD)"
    ws[f'{from_col}3'].font = Font(bold=True)
    ws[f'{from_col}3'].alignment = Alignment(horizontal='center')

# Write scraped info into Excel
def write_to_excel(section_name, product_results, today_str, file_path='CarParts_Pricing.xlsx'):
    base_date = datetime(2025, 4, 24)
    today_date = datetime.strptime(today_str, "%m/%d/%Y")
    days_since = (today_date - base_date).days
    row_index = 4 + days_since

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet.title == "Sheet":
            wb.remove(default_sheet)

    if section_name in wb.sheetnames:
        ws = wb[section_name]
    else:
        ws = wb.create_sheet(section_name)
        ws['A1'] = 'Name'
        ws['A2'] = 'Part # / SKU'
        ws['A3'] = 'Date'
        for r in range(1, 4):
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')

    start_col = 2
    for idx, product in enumerate(product_results):
        name = product["Name"]
        sku = product["SKU"]
        price = product["Price"]
        url = product["URL"]
        col = start_col + idx

        name_cell = ws.cell(row=1, column=col)
        name_cell.value = name
        if url:
            name_cell.hyperlink = url
            name_cell.style = "Hyperlink"

        ws.cell(row=2, column=col).value = sku

        price_cell = ws.cell(row=row_index, column=col)
        price_cell.value = price
        price_cell.number_format = '"$"#,##0.00'

        previous_price_cell = ws.cell(row=row_index - 1, column=col)
        previous_price = previous_price_cell.value

        if previous_price is not None and isinstance(previous_price, (int, float)) and isinstance(price, (int, float)):
            if price > previous_price:
                color = "FF0000"  # red
            elif price < previous_price:
                color = "00B050"  # green
            else:
                color = "000000"  # black
            price_cell.font = Font(color=color)
        else:
            price_cell.font = Font(color="000000")  # default black

    merge_price_label_row(ws, start_col, len(product_results))
    ws.cell(row=row_index, column=1).value = today_date
    ws.cell(row=row_index, column=1).number_format = 'mm/dd/yyyy'

    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            width = len(str(cell_value)) + 2
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width

    try:
        wb.save(file_path)
        logging.info(f"✅ Excel updated for section: {section_name}")
    except PermissionError:
        print("❌ Excel File open — please close the file and try again.")
        logging.error("❌ Excel File open — save failed.")

# Main execution flow
def main():
    backup_input_file()
    today_str = datetime.now(est).strftime("%m/%d/%Y")

    # Let user add a new URL
    print("\n📥 Add a new product link")
    car_model = input("Enter the car model (e.g., E92 M3): ").strip()
    product_url = input("Enter the product URL: ").strip()

    if car_model and product_url:
        with open('input_links.txt', 'a') as f:
            f.write(f"{car_model}|{product_url}\n")
        print("✅ Link added successfully.\n")
    else:
        print("⚠️ Missing car model or URL. Skipping addition.\n")

    # Read all stored links
    if not os.path.exists('input_links.txt'):
        print("❌ No input_links.txt found. Exiting.")
        return

    sections = {}
    with open('input_links.txt', 'r') as f:
        for line in f:
            if '|' in line:
                car_model, url = line.strip().split('|', 1)
                sections.setdefault(car_model, []).append(url)

    for section, urls in sections.items():
        print(f"\n📦 Processing section: {section}")
        product_data = []

        for url in urls:
            name_guess = url.split('/')[-1].replace('-', ' ').capitalize()
            data = get_product_info(url, name_guess)
            if data:
                product_data.append(data)

        if product_data:
            write_to_excel(section, product_data, today_str)

if __name__ == '__main__':
    main()